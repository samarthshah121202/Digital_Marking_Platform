from django.shortcuts import render, redirect, get_object_or_404 # Importing the necessary functions
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from .models import Assignment, StudentWork, Section, StudentMark, Question, Feedback
import os
from .forms import AssignmentForm
from .utils import add_to_feedback_sheet, handle_uploaded_file, handle_upload_excel_sheet, extract_student_info_from_pdf, create_feedback_doc, question_mark_excel, question_mark_excel_student
from django.conf import settings
import csv
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Prefetch
import json
from django.db import transaction
from django.templatetags.static import static
from django.contrib import messages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from django.http import FileResponse
from django.shortcuts import render, get_object_or_404
from django.urls import reverse, path 
import logging 
import re
from django.http import HttpResponse, FileResponse

logger = logging.getLogger(__name__)


filepath_feedback = ""

def get_spreadsheet_name(request, project_name):
    logger.info(f"User {request.user.username} is getting spreadsheet name for project: {project_name}")  # Log for username
    file_name = str(project_name) + "_student_marks.xlsx"
    project_folder = os.path.join(settings.MEDIA_ROOT, "assignments", request.user.username, project_name) 
    return os.path.join(project_folder, file_name)

def create_assignment(request): 
    logger.info(f"User {request.user.username} is attempting to create an assignment.")  # Log for username
    if request.method == "POST": 
        form = AssignmentForm(request.POST, request.FILES) 
        logger.debug(f"Form data received: {form.data}") 
        logger.debug(f"Files received: {request.FILES}") 

        if form.is_valid(): 
            try:
                student_num = []
                group_num = []
                
                project_name = form.cleaned_data["project_name"] 
                project_folder = os.path.join(settings.MEDIA_ROOT, "assignments", request.user.username, project_name) 
                student_submissions_path = os.path.join(project_folder, "student_submissions")
                student_feedback_doc_path =  os.path.join(project_folder, "student_feedback")
                os.makedirs(student_submissions_path,exist_ok=True) 
                os.makedirs(student_feedback_doc_path,exist_ok=True)

                
                assignment = Assignment.objects.create(
                    user=request.user,
                    project_name=project_name, 
                    is_group_assignment=form.cleaned_data["is_group_assignment"]  
                )
                
                markscheme_breakdown = []
                for index, file in enumerate(request.FILES.getlist('student_work')): 
                    
                    student_work_path = os.path.join(student_submissions_path, f"student_work_{index}.pdf") 
                    handle_uploaded_file(file, student_work_path)

                    
                    
                    student_info = extract_student_info_from_pdf(student_work_path, is_group=form.cleaned_data["is_group_assignment"])
                   
                    for student in student_info:
                        if form.cleaned_data["is_group_assignment"]:
                            if student['group_number'] not in group_num:
                                group_num.append(student['group_number'])
                        else:
                            # Append the student number to the student_num list
                             student_num.append(student['student_number'])
                    
                    
                
                    # Print the list of student numbers
                    
                    file_path = os.path.join( request.user.username, project_name, "student_submissions", f"student_work_{index}.pdf")
                    
                    for student in student_info: # Iterate over each student in the student info list
                        # Optionally, save the extracted student info to the StudentWork model
                        student_work = StudentWork.objects.create( # Create a new StudentWork instance
                            student_file_path = file_path,
                            first_name=student['first_name'], # Set the first name
                            last_name=student['last_name'], # Set the last name
                            student_number=student['student_number'], # Set the student number
                            assignment=assignment, # Associate the student work with the created assignment

                            group_number= student['group_number'] if form.cleaned_data["is_group_assignment"] else 1# Associate the student work with the created assignment
                        )
                        # Add the student work instance to the assignment
                        
                            
                          # Corrected line to append student nu
                # Process the mark scheme Excel file if uploaded
                if 'markscheme' in request.FILES: # Check if the mark scheme file is uploaded
                    markscheme_path = os.path.join(project_folder, "markscheme","markscheme.xlsx")
                    handle_uploaded_file(request.FILES["markscheme"], markscheme_path)          
                    # Handle the uploaded Excel mark scheme (conversion to CSV or other processing)
                    sections = handle_upload_excel_sheet(markscheme_path, project_folder, "markscheme", assignment)

                    for section in sections:
                        markscheme_breakdown.append(section.section_name)

                        for module in section.modules.all():
                            markscheme_breakdown.append(module.module_name)

                            for question in module.questions.all():
                                markscheme_breakdown.append(question.question)
                print(markscheme_breakdown)
                # Create and save the marks workbook
                wb = Workbook()
                
                sheet_names = ["Marks Breakdown","Id List", "Group List"] if assignment.is_group_assignment else ["Marks Breakdown","Id List"]
                for idx, name in enumerate(sheet_names):
                    wb.create_sheet(name, idx)
                
                wb.remove(wb.worksheets[-1])

                marks_breakdown_sheet = wb[sheet_names[0]]
                marks_breakdown_sheet.cell(row=1, column=1, value="Criteria")
                               
                for row_num, item in enumerate(markscheme_breakdown, start=2):
                    marks_breakdown_sheet.cell(row=row_num, column=1, value=item)

                row_total = marks_breakdown_sheet.max_row + 1  # Insert total row after the last data row
                marks_breakdown_sheet.cell(row=row_total, column=1, value="Total")  # Add "Total" in column A


                row_grade = marks_breakdown_sheet.max_row + 1  # Insert total row after the last data row
                marks_breakdown_sheet.cell(row=row_grade, column=1, value="Provisional Grade")  # Add "Total" in column A

                col_num = 2
                if assignment.is_group_assignment:
                    for col_num, item in enumerate(group_num, start=2):
                        value_to_write = f"Group {item}" 
                        marks_breakdown_sheet.cell(row=1, column=col_num, value=value_to_write)

                else:
                    for col_num, item in enumerate(student_num, start=2):
                        marks_breakdown_sheet.cell(row=1, column=col_num, value=item)
                
                id_sheet = wb[sheet_names[1]]
                headers = ["Student Id", "Student Name", "Mark"]
                for col_num, header in enumerate(headers, start=1):
                    id_sheet.cell(row=1, column=col_num, value=header)

                # Save the workbook in the project folder
                marks_file_path = os.path.join(project_folder, str(project_name) + "_student_marks.xlsx")
                logger.info(f"path {marks_file_path}")
                wb.save(marks_file_path)

                # Save the assignment instance after creating student works and handling files
                assignment.save()

                logger.info(f"User {request.user.username} successfully created assignment: {assignment.id}")  # Log successful creation
                return redirect('assignment_detail', assignment_id=assignment.id)  # Redirect to the dashboard or any appropriate page

            except Exception as e:
                logger.error(f"Error during file upload and processing for user {request.user.username}: {str(e)}")  
                return render(request, 'main/create_assignment.html', { # Render the create assignment template
                    'form': form, # Pass the form to the template
                    'error': f"Error uploading files: {str(e)}"
                })
        else:
            logger.info(f"Form validation errors for user {request.user.username}: {form.errors}")  
            return render(request, 'main/create_assignment.html', {
                'form': form,
                'error': "Form validation failed. Please check your inputs."
            })
    else:
        form = AssignmentForm()
    return render(request, 'main/create_assignment.html', {'form': form})

def homepage(request):
    logger.info(f"User {request.user.username} accessed the homepage.")  # Log for username
    return render(request, 'main/homepage.html')

def register(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        
        # Check if the username already exists
        if User.objects.filter(username=username).exists():
            # Display an error message if the username is taken
            return render(request, 'main/register.html', {'error': 'Username already taken'})
        
        # Create a new user if the username is unique
        user = User.objects.create_user(username=username, password=password)
        login(request, user)
        logger.info(f"User {request.user.username} registered successfully.")  # Log successful registration
        return redirect('dashboard')
    
    return render(request, 'main/register.html')

def login_view(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            logger.info(f"User {username} logged in successfully.")  # Log successful login
            login(request, user)
            return redirect('dashboard')
    return render(request, 'main/login.html')

def logout_view(request):
    logger.info(f"User {request.user.username} logged out.")  # Log for username
    logout(request)
    return redirect('homepage')

@login_required
def dashboard(request): 
    logger.info(f"User {request.user.username} accessed the dashboard.")  # Log for username
    # Filter assignments for the currently logged-in user
    assignments = Assignment.objects.filter(user=request.user)

    # Render the dashboard template with the assignments context
    return render(request, 'main/dashboard.html', {
        'assignments': assignments,
        'user': request.user,  # Optional, in case you want to greet the user or display the username
    })

@login_required
def assignment_detail(request, assignment_id):
    logger.info(f"User {request.user.username} is viewing assignment details for assignment ID: {assignment_id}")  # Log for username
    assignment = Assignment.objects.get(id=assignment_id)
    student_info = StudentWork.objects.filter(assignment=assignment).values(
        'first_name',
        'last_name',
        'student_number',
        'group_number',
        'id',
        "is_marked"
    )

    all_marked = all(student['is_marked'] for student in student_info)

    print(f"All submissions marked: {all_marked}")  

    context = {
        'assignment': assignment,
        'student_info': student_info,
        'all_marked': all_marked
    }

    if assignment.is_group_assignment:
       return render(request, 'main/assignment_detail_group.html', context) 

    return render(request, 'main/assignment_detail_individual.html', context) 
    
@login_required
def delete_assignment(request, assignment_id):
    logger.info(f"User {request.user.username} is attempting to delete assignment: {assignment_id}")  # Log for username
    try:
        # Get the assignment and verify ownership
        assignment = get_object_or_404(Assignment, id=assignment_id, user=request.user)
        
        # Get the project folder path
        project_folder = os.path.join(settings.MEDIA_ROOT, "assignments", 
                                    request.user.username, assignment.project_name)
        
        # Delete all related database records
        assignment.delete()
        
        # Delete the project folder and all its contents
        if os.path.exists(project_folder):
            import shutil
            shutil.rmtree(project_folder)
        
        messages.success(request, 'Assignment deleted successfully')
        logger.info(f"User {request.user.username} successfully deleted assignment: {assignment_id}")  # Log successful deletion
        logger.add_metric('assignment_deleted', {'username': request.user.username, 'assignment_id': assignment_id})  # Log metric
        return redirect('dashboard')
    except Assignment.DoesNotExist:
        messages.error(request, 'Assignment not found')
        return redirect('dashboard')
    except Exception as e:
        messages.error(request, f'Error deleting assignment: {str(e)}')
        return redirect('dashboard')

def extract_marks(question_text):
    match = re.search(r"\((\d+)\smarks?\)", question_text)
    return int(match.group(1)) if match else None

@login_required
def view_markscheme(request, assignment_id, submission_id):
    logger.info(f"User {request.user.username} is viewing markscheme for assignment ID: {assignment_id}, submission ID: {submission_id}")  # Log for username
    try:
        assignment = Assignment.objects.get(id=assignment_id)
        # Get the specific student work
        student_work = assignment.student_works.get(id=submission_id) if submission_id else assignment.student_works.first()
        
        path = student_work.student_file_path
        student_work_path = static(path)
        
        # Debug prints
       # print(f"Student work path: {student_work_path}")
        
        sections = Section.objects.prefetch_related(
            'modules',
            'modules__questions',
            'modules__questions__feedbacks',
        ).filter(assignment=assignment)

        return render(request, 'main/view_markscheme.html', {
            'assignment': assignment,
            'sections': sections,
            'student_work': student_work,
            'student_work_path': student_work_path
        })
    except Assignment.DoesNotExist:
        return render(request, 'main/view_markscheme.html', {
            'error': 'Assignment not found'
        })

@login_required
@require_POST
def save_marks(request):
    logger.info(f"User {request.user.username} is saving marks.")  # Log for username
    def get_students(is_group_project, submission_id):
        if is_group_project:
            first_student_in_group = StudentWork.objects.get(id=submission_id)
            group_number = first_student_in_group.group_number
            return StudentWork.objects.filter(group_number=group_number)

        return StudentWork.objects.get(id=submission_id) 

    try:
        data = json.loads(request.body)
        marks_data = data.get('marks', {})
        submission_id= request.GET.get('submission_id', -1)
        is_group = int(request.GET.get("is_group", "0"))

        student = get_students(is_group == 1, submission_id)
        
        # Start a transaction to ensure all marks are saved or none are
        with transaction.atomic():
            for question_id, mark_info in marks_data.items():
                clean_question_id = question_id.replace('question-', '')
                # Get the question
                question = Question.objects.get(id=clean_question_id)
                
                # Get the feedback and create new mark
                feedback = None
                if 'feedbackId' in mark_info:
                    feedback = Feedback.objects.get(id=mark_info['feedbackId'])

                if is_group == 0:
                    if "customFeedback" in mark_info:
                        StudentMark.objects.create(
                            student=student,
                            question=question,
                            custom_feedback=mark_info.get("customFeedback", ""),
                            custom_mark=mark_info.get("customMark", 0)
                        )
                    else:
                        StudentMark.objects.create(
                            student=student,
                            question=question,
                            feedback=feedback
                        )
                    student.is_marked = True
                    student.save()
                else:
                    for participant in student:
                        if "customFeedback" in mark_info:
                            StudentMark.objects.create(
                                student=participant,
                                question=question,
                                custom_feedback=mark_info.get("customFeedback", ""),
                                custom_mark=mark_info.get("customMark", 0)
                            )
                        else:
                            StudentMark.objects.create(
                                student=participant,
                                question=question,
                                feedback=feedback
                            )
                        participant.is_marked = True
                        participant.save()


        return JsonResponse({
            'success': True,
            'message': 'All marks saved successfully'
        })
            
    except Question.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Question not found'
        }, status=404)
    except Feedback.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Feedback not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

@login_required
def view_marks(request, assignment_id, submission_id):
    logger.info(f"User {request.user.username} is viewing marks for assignment ID: {assignment_id}, submission ID: {submission_id}")  # Log for username
    try:
        assignment = Assignment.objects.get(id=assignment_id)

        project_name = assignment.project_name
        project_folder = os.path.join(settings.MEDIA_ROOT, "assignments", request.user.username, project_name)
        student_feedback_doc_path = os.path.join(project_folder, "student_feedback")

        student_work = assignment.student_works.get(id=submission_id)
        feedback_doc_name=""

        sections = Section.objects.prefetch_related(
            'modules',
            'modules__questions',
            'modules__questions__feedbacks',
            'modules__questions__studentmark_set'
        ).filter(
            assignment=assignment,
            modules__questions__studentmark__student=student_work
        ).distinct()

        processed_sections = []
        total_marks = 0

        for section in sections:
            section_total = 0
            processed_modules = []

            for module in section.modules.all():
                module_total = 0
                processed_questions = []

                for question in module.questions.all():
                    student_mark = question.studentmark_set.filter(student=student_work).first()

                    question_data = {
                        'question': question,
                        'mark': student_mark.feedback.mark if student_mark and student_mark.feedback else student_mark.custom_mark,
                        'feedback_text': student_mark.feedback.feedback_text if student_mark and student_mark.feedback else "",
                        'custom_feedback': student_mark.custom_feedback if student_mark else "No custom feedback recorded",
                    }
                    processed_questions.append(question_data)

                    if student_mark:
                        if student_mark.feedback:
                            module_total += student_mark.feedback.mark
                        else:
                            module_total += student_mark.custom_mark

                module_data = {
                    'module': module,
                    'questions': processed_questions,
                    'total': module_total
                }

                processed_modules.append(module_data)
                section_total += module_total

            section_data = {
                'section': section,
                'modules': processed_modules,
                'total': section_total
            }
            processed_sections.append(section_data)
            
            total_marks += section_total
        print(processed_sections)
        feedback_above_50 = []
        feedback_below_50 = []

        for section in processed_sections:
            for module in section['modules']:
                for question in module['questions']:
                    mark = question['mark']
                    feedback_text = question['feedback_text']
                    
                    question_text = str(question)
                    match = re.search(r'\((\d+) marks\)', question_text)

                    if match:
                        max_marks = float(match.group(1))  

                    if max_marks is not None:
                        if mark >= 0.5 * max_marks:
                            feedback_above_50.append(feedback_text)
                        else:
                            feedback_below_50.append(feedback_text)
                    else:
                        print(f"Skipping question due to missing max_marks: {question['question']}")



        marks_file_path = os.path.join(project_folder, str(project_name) + "_student_marks.xlsx")
        wb = load_workbook(marks_file_path)
        
        if assignment.is_group_assignment:
            question_mark_excel(wb, processed_questions, processed_modules, processed_sections, student_work.group_number, total_marks)
        else:
            question_mark_excel_student(wb, processed_questions, processed_modules, processed_sections, student_work.student_number, total_marks)


        wb.save(marks_file_path)
        wb.close()
        #print("SAVED AND CLOSED)")
        # Modify the headers based on assignment type
        if assignment.is_group_assignment:
            student_info = [["First Name", "Last Name", "ID", "Group No."]]
            group_members = StudentWork.objects.filter(group_number=student_work.group_number, assignment=assignment).distinct()
            id_table = []
            group_table = [["Group Number:" + str(group_members[0].group_number), "Group Mark:" + str(total_marks)],
                           ["Name", "Student ID", "Individual Mark"]]
            for member in group_members:
                student_info.append([
                    member.first_name, 
                    member.last_name, 
                    member.student_number,
                    member.group_number  # Add group number to each student's info
                ])
                id_table.append([member.student_number, member.first_name + " " + member.last_name, total_marks])
                group_table.append([member.first_name + " " + member.last_name, member.student_number,total_marks]) 

        else:
            student_info = [["First Name", "Last Name", "ID"]]
            tmp = StudentWork.objects.get(id=submission_id)
            student_info.append([tmp.first_name, tmp.last_name, tmp.student_number])
            id_table = [[tmp.student_number, tmp.first_name + " " + tmp.last_name, total_marks]]

        marks_file_path = os.path.join(project_folder, str(project_name) + "_student_marks.xlsx")
        wb = load_workbook(marks_file_path)


        feedback_doc_path = static(create_feedback_doc(student_info,processed_sections, assignment.project_name, student_feedback_doc_path, assignment, student_work, feedback_above_50, feedback_below_50).removeprefix("assignments"))
        logger.info(f"Feedback path = {feedback_doc_path}")
        
        if assignment.is_group_assignment:
            add_to_feedback_sheet(wb, id_table, group_table)
            feedback_doc_name = f"Group_{student_work.group_number}_Feedback"
        else:
            add_to_feedback_sheet(wb, id_table)
            feedback_doc_name = f"{tmp.student_number}_Feedback"


        wb.save(marks_file_path)
        wb.close()

     

        return render(request, 'main/view_marks.html', {
            'assignment': assignment,
            'processed_sections': processed_sections,
            'total_marks': total_marks,
            'submission_id': student_work.id,
            'feedback_doc_path': feedback_doc_path,
            'feedback_doc_name': feedback_doc_name
        })
        
    except Assignment.DoesNotExist:
        return render(request, 'main/view_marks.html', {
            'error': 'Assignment not found'
        })


def finish_assignment(request, assignment_id):
    logger.info(f"User {request.user.username} is finishing assignment ID: {assignment_id}")  # Log for username
    # Logic for finishing the assignment (e.g., marking it as complete)
    assignment = get_object_or_404(Assignment, id=assignment_id, user=request.user)
    excel_name = ""
    spreadsheet_path = get_spreadsheet_name(request=request, project_name=assignment.project_name)
    if spreadsheet_path and spreadsheet_path.startswith("assignments"):
        spreadsheet_path = spreadsheet_path.removeprefix("assignments")

    static_spreadsheet_path = static(spreadsheet_path)
    excel_name = "{assignment.project_name}__student_marks"
    print(static_spreadsheet_path)


    return render(request, 'main/finish_assignment.html', {
        'assignment': assignment,
        'static_spreadsheet_path': static_spreadsheet_path,
        'excel_name': excel_name
    })